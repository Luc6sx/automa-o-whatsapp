import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import PySimpleGUI as sg
import xlsxwriter
from datetime import datetime

# Interface
sg.theme('Reddit')
layout = [
    [sg.Text('Escolha a planilha para enviar as mensagens')],
    [sg.Input(), sg.FileBrowse('Escolher Planilha', file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Text('Nova data de vencimento (formato: DD/MM/YYYY)')],
    [sg.Input(key='-NOVA_DATA-', size=(20, 1))],
    [sg.Button('Planilha Atualizada', size=(20, 1))],
    [sg.Button('Enviar Mensagens', size=(20, 1))],
    [sg.Button('Parar', size=(20,1))],
    [sg.Output(size=(50, 10))]  # Área para mostrar os logs
]

# Janela
janela = sg.Window('Bot para Enviar Mensagens', layout)

# Eventos
while True:
    evento, valores = janela.read()

    if evento == sg.WINDOW_CLOSED or evento == 'Parar':
        print("programa Encerrado")
        break

    if evento == 'Enviar Mensagens':
        arquivo_planilha = valores[0]
        if not arquivo_planilha:
            print("Por favor, escolha uma planilha.")
            continue

        try:
            # Carregar a planilha
            workbook = openpyxl.load_workbook(arquivo_planilha)
            pagina_clientes = workbook.active  # Seleciona a primeira planilha
            
            for linha in pagina_clientes.iter_rows(min_row=2):
                nome = linha[0].value
                telefone = linha[1].value
                vencimento = linha[2].value

                # Formatar a mensagem
                mensagem = f'Olá {nome}, seu boleto vence no dia {vencimento.strftime("%d/%m/%Y")}. Favor pagar no link: https://www.link_do_pagamento.com'
                link_mensagem_whats = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
                
                # Abrir o WhatsApp Web e enviar a mensagem
                webbrowser.open(link_mensagem_whats)
                sleep(20)  # Ajuste o tempo conforme necessário

                try:
                    # Enviar a mensagem pressionando Enter
                    pyautogui.press('enter')
                    sleep(3)
                    pyautogui.hotkey('ctrl', 'w')  # Fecha a aba do WhatsApp Web
                    sleep(3)

                    print(f'Mensagem enviada para {nome} ({telefone})')
                except Exception as e:
                    print(f'Não foi possível enviar mensagem para {nome} ({telefone}): {e}')
                    with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
                        arquivo.write(f'{nome},{telefone}\n')

        except Exception as e:
            print(f'Erro ao processar a planilha: {e}')

    if evento == 'Planilha Atualizada':
        arquivo_planilha = valores[0]
        nova_data_str = valores['-NOVA_DATA-']

        if not arquivo_planilha:
            print("Por favor, escolha uma planilha.")
            continue

        if not nova_data_str:
            print("Por favor, insira a nova data de vencimento.")
            continue

        try:
            # Converter a nova data para o formato datetime
            nova_data = datetime.strptime(nova_data_str, "%d/%m/%Y")
        except ValueError:
            print("Formato de data inválido. Use DD/MM/YYYY.")
            continue

        try:
            # Carregar a planilha
            workbook = openpyxl.load_workbook(arquivo_planilha)
            pagina_clientes = workbook.active  # Seleciona a primeira planilha

            

            # Alterar o vencimento e salvar na nova planilha
            for index, linha in enumerate(pagina_clientes.iter_rows(min_row=2), start=2):
                nome = linha[0].value
                telefone = linha[1].value
                linha[2].value = nova_data  # Substituir o vencimento pela nova data

                print(f'Vencimento de {nome} atualizado para {nova_data.strftime("%d/%m/%Y")}')

            # Salvar o arquivo original atualizado
            workbook.save(arquivo_planilha)
            
            print("Planilha atualizada criada com sucesso.")
        except Exception as e:
            print(f'Erro ao processar a planilha: {e}')


janela.close()
